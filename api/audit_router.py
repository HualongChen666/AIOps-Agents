# -*- coding: utf-8 -*-
"""Audit Export & Report API

提供审计日志的 CSV / Excel 导出以及基于审计日志的统计报告（JSON）
仅在提供正确的 `X-Internal-Key`（或未配置 INTERNAL_API_KEY）时可访问。
"""

import csv
import os
import tempfile
from typing import Any, Dict, List

from fastapi import APIRouter, BackgroundTasks, HTTPException, Query, Request
from fastapi.responses import FileResponse, JSONResponse
from loguru import logger

from config import INTERNAL_API_KEY  # Internal API key for protected routes
from core.command_guard import get_audit_log
from core.compliance import mask_sensitive_dict

router = APIRouter(prefix="/api/v1/audit", tags=["Audit Export & Report"])


def _safe_remove_file(file_path: str) -> None:
    """Safely remove a file with error handling."""
    try:
        os.remove(file_path)
    except OSError as exc:
        logger.warning(f"Failed to remove temporary file {file_path}: {exc}")


def _get_and_mask_audit_logs(limit: int) -> List[Dict[str, Any]]:
    """
    获取并脱敏审计日志

    Args:
        limit: 获取日志的数量限制

    Returns:
        脱敏后的审计日志列表
    """
    raw_logs: List[Dict[str, Any]] = get_audit_log(limit=limit)
    return [mask_sensitive_dict(log) for log in raw_logs]


def _write_empty_csv_file(tmp_path: str, header_fields: List[str]) -> None:
    """
    写入空CSV文件

    Args:
        tmp_path: 临时文件路径
        header_fields: 表头字段列表

    Raises:
        HTTPException: 写入失败
    """
    try:
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=header_fields)
            writer.writeheader()
    except (OSError, IOError) as exc:
        logger.error(f"Failed to write empty CSV file: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to generate empty CSV export")


def _write_empty_excel_file(tmp_path: str, header_fields: List[str]) -> None:
    """
    写入空Excel文件

    Args:
        tmp_path: 临时文件路径
        header_fields: 表头字段列表

    Raises:
        HTTPException: 写入失败或openpyxl未安装
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"openpyxl 未安装: {e}")

    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Audit Log")
        ws.append(header_fields)
        wb.save(tmp_path)
    except Exception as exc:
        logger.error(f"Failed to write empty Excel file: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to generate empty Excel export")


def _create_empty_export_file(fmt: str, background_tasks: BackgroundTasks) -> FileResponse:
    """
    创建空导出文件

    Args:
        fmt: 导出格式 (csv 或 excel)
        background_tasks: 后台任务对象

    Returns:
        文件响应对象

    Raises:
        HTTPException: 创建失败
    """
    header_fields = ["timestamp", "event", "risk_level", "result"]

    if fmt == "csv":
        fd, tmp_path = tempfile.mkstemp(suffix=".csv")
        os.close(fd)
        _write_empty_csv_file(tmp_path, header_fields)
        filename = f"audit_export_{fmt}.csv"
        background_tasks.add_task(_safe_remove_file, tmp_path)
        return FileResponse(
            path=tmp_path, filename=filename, media_type="text/csv", background=background_tasks
        )
    else:
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
        os.close(fd)
        _write_empty_excel_file(tmp_path, header_fields)
        filename = f"audit_export_{fmt}.xlsx"
        background_tasks.add_task(_safe_remove_file, tmp_path)
        return FileResponse(
            path=tmp_path,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=background_tasks,
        )


def _write_csv_file(tmp_path: str, logs: List[Dict[str, Any]]) -> None:
    """
    写入CSV文件

    Args:
        tmp_path: 临时文件路径
        logs: 审计日志列表

    Raises:
        HTTPException: 写入失败
    """
    try:
        with open(tmp_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=logs[0].keys())
            writer.writeheader()
            writer.writerows(logs)
    except (OSError, IOError) as exc:
        logger.error(f"Failed to write CSV file: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to generate audit export")


def _write_excel_file(tmp_path: str, logs: List[Dict[str, Any]]) -> None:
    """
    写入Excel文件

    Args:
        tmp_path: 临时文件路径
        logs: 审计日志列表

    Raises:
        HTTPException: 写入失败或openpyxl未安装
    """
    try:
        from openpyxl import Workbook
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f"openpyxl 未安装: {e}")

    try:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Audit Log")
        ws.append(list(logs[0].keys()))
        for row in logs:
            ws.append([row.get(col) for col in logs[0].keys()])
        wb.save(tmp_path)
    except Exception as exc:
        logger.error(f"Failed to write Excel file: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to generate audit export")


def _create_export_file(
    fmt: str, logs: List[Dict[str, Any]], background_tasks: BackgroundTasks
) -> FileResponse:
    """
    创建导出文件

    Args:
        fmt: 导出格式 (csv 或 excel)
        logs: 审计日志列表
        background_tasks: 后台任务对象

    Returns:
        文件响应对象

    Raises:
        HTTPException: 创建失败
    """
    suffix = ".csv" if fmt == "csv" else ".xlsx"
    fd, tmp_path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)

    try:
        if fmt == "csv":
            _write_csv_file(tmp_path, logs)
        else:
            _write_excel_file(tmp_path, logs)
    except Exception as exc:
        logger.error(f"Failed to generate audit export: {exc}")
        raise HTTPException(status_code=500, detail=f"Failed to generate audit export")

    filename = f"audit_export_{fmt}{suffix}"
    background_tasks.add_task(_safe_remove_file, tmp_path)
    return FileResponse(
        path=tmp_path,
        filename=filename,
        media_type="application/octet-stream",
        background=background_tasks,
    )


def _verify_internal_key(request: Request) -> None:
    """验证内部 API Key，仅本地或通过环境变量配置的请求可访问。
    当 `INTERNAL_API_KEY` 为空时，视为未启用校验，直接通过。
    """
    if not INTERNAL_API_KEY:
        # 未设置内部密钥，直接放行（仅在开发/内部环境下使用）
        return
    provided_key = request.headers.get("X-Internal-Key")
    if not provided_key:
        raise HTTPException(status_code=403, detail="Missing X-Internal-Key header")
    if provided_key != INTERNAL_API_KEY:
        raise HTTPException(status_code=403, detail="Invalid X-Internal-Key")


@router.get(
    "/export",
    summary="导出审计日志",
    description="以 CSV 或 Excel 格式导出审计日志",
    responses={
        200: {"description": "审计日志文件(CSV或Excel)"},
        403: {"description": "权限不足(需要X-Internal-Key)"},
        500: {"description": "导出失败"},
    },
)
# logger imported at top
async def export_audit(
    request: Request,
    background_tasks: BackgroundTasks,
    fmt: str = Query("csv", pattern="^(csv|excel)$", description="导出格式: csv 或 excel"),
    limit: int = Query(100, ge=1, le=5000, description="导出记录数量上限"),
) -> FileResponse:
    """导出审计日志为 CSV 或 Excel 文件。

    访问需提供 `X-Internal-Key` 头部进行身份校验。
    """
    _verify_internal_key(request)
    logger.debug("export_audit called with fmt=%s limit=%d", fmt, limit)

    # 获取并脱敏审计日志
    logs = _get_and_mask_audit_logs(limit)

    # 如果没有日志，返回空文件
    if not logs:
        return _create_empty_export_file(fmt, background_tasks)

    # 创建导出文件
    return _create_export_file(fmt, logs, background_tasks)


@router.get(
    "/report",
    summary="生成审计报告",
    description="基于审计日志生成简易报告（JSON）",
    responses={
        200: {
            "description": "审计报告(JSON)",
            "content": {
                "application/json": {
                    "example": {
                        "total": 100,
                        "risk_distribution": {"high": 10, "medium": 30, "low": 50, "safe": 10},
                        "result_distribution": {"blocked": 5, "allowed": 95},
                        "sample": [],
                    }
                }
            },
        },
        403: {"description": "权限不足(需要X-Internal-Key)"},
        500: {"description": "生成报告失败"},
    },
)
async def audit_report(
    request: Request,
    limit: int = Query(100, ge=1, le=5000, description="报告包含的记录数量上限"),
) -> JSONResponse:
    """返回审计日志的聚合统计报告（JSON）。

    仅返回统计信息，例如不同风险等级的计数、最近的错误等。
    """
    _verify_internal_key(request)
    logs = get_audit_log(limit=limit)
    if not logs:
        report = {
            "total": 0,
            "risk_distribution": {},
            "result_distribution": {},
            "sample": [],
        }
        return JSONResponse(content=report)

    from collections import Counter

    risk_counter: Counter[str] = Counter()
    result_counter: Counter[str] = Counter()
    for entry in logs:
        risk_counter[entry.get("risk_level", "UNKNOWN")] += 1
        result_counter[entry.get("result", "UNKNOWN")] += 1

    report = {
        "total": len(logs),
        "risk_distribution": dict(risk_counter),
        "result_distribution": dict(result_counter),
        "sample": logs[:5],
    }
    return JSONResponse(content=report)


@router.get(
    "",
    summary="获取审计日志",
    description="返回最近的审计日志记录",
    responses={
        200: {"description": "审计日志列表(JSON)"},
        403: {"description": "权限不足(需要X-Internal-Key)"},
    },
)
async def list_audit(
    request: Request,
    limit: int = Query(100, ge=1, le=5000, description="返回记录数量上限"),
) -> JSONResponse:
    """返回最近的审计日志（JSON 列表）。"""
    _verify_internal_key(request)
    logs = get_audit_log(limit=limit)
    return JSONResponse(content=logs)
